from flask import current_app
from models.reportes_model import reportes
from utils.search_builder import SearchBuilder

def listarReportes(page=1, limit=50, q=None, order_by=None, **filters):
    c = current_app.mysql.connection.cursor()
    sb = SearchBuilder(
        table='t_reporte',
        search_fields=['rep_id', 'rep_tipo', 'rep_parametros'],
        exact_fields=['rep_tipo', 'rep_usu_id_fk'],
        range_fields={'rep_fecha': 'date'},
        default_order='rep_fecha DESC'
    )
    result = sb.execute(c, page=page, limit=limit, q=q, order_by=order_by, **filters)
    c.close()

    lista = []
    for item in result['data']:
        r = reportes(item['rep_id'], item['rep_tipo'], item['rep_fecha'],
                     item['rep_parametros'], item['rep_usu_id_fk'], item.get('rep_resultado')).todic()
        lista.append(r)

    result['data'] = lista
    return result

def registrarReportes(REP_ID, REP_TIPO, REP_FECHA, REP_PARAMETROS, REP_USU_ID_FK, REP_RESULTADO):
    c = current_app.mysql.connection.cursor()
    sql = """
        INSERT INTO t_reporte (rep_id, rep_tipo, rep_fecha, rep_parametros, rep_usu_id_fk, rep_resultado) 
        VALUES (%s, %s, %s, %s, %s, %s)
    """
    c.execute(sql, (REP_ID, REP_TIPO, REP_FECHA, REP_PARAMETROS, REP_USU_ID_FK, REP_RESULTADO))
    current_app.mysql.connection.commit()
    id = c.lastrowid
    c.close()
    return reportes(REP_ID, REP_TIPO, REP_FECHA, REP_PARAMETROS, REP_USU_ID_FK, REP_RESULTADO).todic()



def eliminarReportes(REP_ID):
    c = current_app.mysql.connection.cursor()
    sql = "DELETE FROM t_reporte WHERE rep_id = %s"
    c.execute(sql, (REP_ID,))
    current_app.mysql.connection.commit()
    c.close()
    return {"mensaje": "Reporte eliminado", "rep_id": REP_ID}

def buscarReportes(REP_ID): # <-- IMPORTANTE: Debe tener REP_ID
    c = current_app.mysql.connection.cursor()
    sql = """
        SELECT rep_id, rep_tipo, rep_fecha, rep_parametros, rep_usu_id_fk, rep_resultado 
        FROM t_reporte 
        WHERE rep_id = %s
    """
    c.execute(sql, (REP_ID,))
    p = c.fetchone()
    c.close()
    
    if p:
        return reportes(p[0], p[1], p[2], p[3], p[4], p[5]).todic()
    return None

def editarReportes(REP_ID, REP_TIPO, REP_FECHA, REP_PARAMETROS, REP_USU_ID_FK, REP_RESULTADO):
    c = current_app.mysql.connection.cursor()
    sql = """
        UPDATE t_reporte 
        SET rep_tipo=%s, rep_fecha=%s, rep_parametros=%s, rep_usu_id_fk=%s, rep_resultado=%s
        WHERE rep_id=%s
    """
    c.execute(sql, (REP_TIPO, REP_FECHA, REP_PARAMETROS, REP_USU_ID_FK, REP_RESULTADO, REP_ID))
    current_app.mysql.connection.commit()
    c.close()
    return reportes(REP_ID, REP_TIPO, REP_FECHA, REP_PARAMETROS, REP_USU_ID_FK, REP_RESULTADO).todic()


# ═══════════════════════════════════════════════════════════════════
#  GENERACIÓN REAL DE REPORTES (agregaciones de negocio)
# ═══════════════════════════════════════════════════════════════════

def reporte_ventas(fecha_desde=None, fecha_hasta=None):
    """
    Reporte de ventas: total de pedidos por día/período.
    Retorna lista de {fecha, cantidad_pedidos, total_vendido}.
    """
    c = current_app.mysql.connection.cursor()
    params = []
    where = "WHERE p.ped_estado_entrega != 'Anulado'"
    if fecha_desde:
        where += " AND p.ped_fecha >= %s"
        params.append(fecha_desde)
    if fecha_hasta:
        where += " AND p.ped_fecha <= %s"
        params.append(fecha_hasta)

    c.execute(f"""
        SELECT p.ped_fecha, COUNT(*) AS cantidad, SUM(p.ped_total) AS total
        FROM t_pedido p
        {where}
        GROUP BY p.ped_fecha
        ORDER BY p.ped_fecha DESC
    """, tuple(params))

    resultados = [{"fecha": str(r[0]), "cantidad_pedidos": r[1], "total_vendido": float(r[2] or 0)}
                  for r in c.fetchall()]
    c.close()
    return resultados


def reporte_inventario():
    """
    Reporte de inventario actual: productos con stock, lote, vencimiento.
    """
    c = current_app.mysql.connection.cursor()
    c.execute("""
        SELECT p.pro_id, p.pro_nombre, p.pro_categoria, p.pro_cantidad_disponible,
               p.pro_stock_minimo, p.pro_precio,
               l.lot_id, l.lot_numero, l.lot_fecha_vencimiento, l.lot_cantidad_actual,
               l.lot_estado
        FROM t_producto p
        LEFT JOIN t_lote l ON l.lot_pro_id_fk = p.pro_id
        WHERE p.pro_estado = 'Activo'
        ORDER BY p.pro_nombre ASC, l.lot_fecha_vencimiento ASC
    """)
    resultados = [{
        "pro_id": r[0], "pro_nombre": r[1], "pro_categoria": r[2],
        "pro_cantidad_disponible": r[3], "pro_stock_minimo": r[4],
        "pro_precio": float(r[5] or 0),
        "lot_id": r[6], "lot_numero": r[7],
        "lot_fecha_vencimiento": str(r[8]) if r[8] else None,
        "lot_cantidad_actual": r[9], "lot_estado": r[10]
    } for r in c.fetchall()]
    c.close()
    return resultados


def reporte_mas_vendidos(fecha_desde=None, fecha_hasta=None, limite=10):
    """
    Productos más vendidos en un período, con filtros de fecha reales.
    """
    c = current_app.mysql.connection.cursor()
    params = []
    where = "WHERE p.ped_estado_entrega != 'Anulado'"
    if fecha_desde:
        where += " AND p.ped_fecha >= %s"
        params.append(fecha_desde)
    if fecha_hasta:
        where += " AND p.ped_fecha <= %s"
        params.append(fecha_hasta)

    c.execute(f"""
        SELECT pr.pro_id, pr.pro_nombre,
               SUM(dp.det_cantidad) AS total_unidades,
               SUM(dp.det_subtotal) AS total_vendido
        FROM t_detalle_pedido dp
        JOIN t_pedido p ON p.ped_id = dp.det_ped_id_fk
        JOIN t_producto pr ON pr.pro_id = dp.det_pro_id_fk
        {where}
        GROUP BY pr.pro_id, pr.pro_nombre
        ORDER BY total_vendido DESC
        LIMIT %s
    """, tuple(params + [limite]))

    resultados = [{
        "pro_id": r[0], "nombre": r[1],
        "total_unidades": int(r[2] or 0),
        "total_vendido": float(r[3] or 0)
    } for r in c.fetchall()]
    c.close()
    return resultados


def reporte_por_vencer(dias=30):
    """
    Medicamentos por vencer: lotes activos que vencen dentro de N días.
    """
    c = current_app.mysql.connection.cursor()
    c.execute("""
        SELECT p.pro_id, p.pro_nombre, p.pro_categoria,
               l.lot_id, l.lot_numero, l.lot_fecha_vencimiento,
               l.lot_cantidad_actual,
               DATEDIFF(l.lot_fecha_vencimiento, CURDATE()) AS dias_restantes
        FROM t_lote l
        JOIN t_producto p ON p.pro_id = l.lot_pro_id_fk
        WHERE l.lot_estado = 'Activo'
          AND l.lot_fecha_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)
          AND l.lot_fecha_vencimiento > CURDATE()
        ORDER BY l.lot_fecha_vencimiento ASC
    """, (dias,))

    resultados = [{
        "pro_id": r[0], "pro_nombre": r[1], "pro_categoria": r[2],
        "lot_id": r[3], "lot_numero": r[4],
        "lot_fecha_vencimiento": str(r[5]) if r[5] else None,
        "lot_cantidad_actual": r[6], "dias_restantes": r[7]
    } for r in c.fetchall()]
    c.close()
    return resultados


# ═══════════════════════════════════════════════════════════════════
#  EXPORTACIÓN PDF y EXCEL
# ═══════════════════════════════════════════════════════════════════

from io import BytesIO
from datetime import date as date_cls


def _generar_pdf(titulo, columnas, filas):
    """Genera un PDF simple con reportlab y retorna bytes."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(titulo, styles['Title']))
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(
        f"Generado: {date_cls.today().strftime('%d/%m/%Y')} — San Diego Distribuidora",
        styles['Normal']
    ))
    elementos.append(Spacer(1, 12))

    data = [columnas] + filas
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    return buf.getvalue()


def _generar_excel(titulo, columnas, filas):
    """Genera un archivo Excel (.xlsx) con openpyxl y retorna bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Encabezado
    header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Datos
    for row_idx, fila in enumerate(filas, 2):
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = Font(name='Arial', size=9)
            cell.border = thin_border
            if isinstance(valor, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Ajustar anchos
    for col_idx in range(1, len(columnas) + 1):
        max_len = len(str(columnas[col_idx - 1]))
        for row_idx in range(2, len(filas) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_reporte_pdf(tipo, fecha_desde=None, fecha_hasta=None):
    """Genera y retorna bytes del PDF según el tipo de reporte."""
    if tipo == 'ventas':
        datos = reporte_ventas(fecha_desde, fecha_hasta)
        col = ['Fecha', 'Cantidad Pedidos', 'Total Vendido']
        filas = [[d['fecha'], d['cantidad_pedidos'], f"${d['total_vendido']:,.0f}"] for d in datos]
        titulo = 'Reporte de Ventas'
    elif tipo == 'inventario':
        datos = reporte_inventario()
        col = ['Producto', 'Categoría', 'Stock', 'Stock Mín', 'Precio', 'Lote', 'Vencimiento', 'Cant. Lote']
        filas = [[d['pro_nombre'], d['pro_categoria'], d['pro_cantidad_disponible'],
                  d['pro_stock_minimo'], f"${d['pro_precio']:,.0f}", d['lot_numero'] or '-',
                  d['lot_fecha_vencimiento'] or '-', d['lot_cantidad_actual'] or 0] for d in datos]
        titulo = 'Reporte de Inventario'
    elif tipo == 'mas_vendidos':
        datos = reporte_mas_vendidos(fecha_desde, fecha_hasta)
        col = ['Producto', 'Unidades Vendidas', 'Total Vendido']
        filas = [[d['nombre'], d['total_unidades'], f"${d['total_vendido']:,.0f}"] for d in datos]
        titulo = 'Productos Más Vendidos'
    elif tipo == 'por_vencer':
        datos = reporte_por_vencer()
        col = ['Producto', 'Categoría', 'Lote', 'Vencimiento', 'Stock Lote', 'Días Rest.']
        filas = [[d['pro_nombre'], d['pro_categoria'], d['lot_numero'] or d['lot_id'],
                  d['lot_fecha_vencimiento'], d['lot_cantidad_actual'], d['dias_restantes']] for d in datos]
        titulo = 'Medicamentos por Vencer'
    else:
        return None

    return _generar_pdf(titulo, col, filas)


def exportar_reporte_excel(tipo, fecha_desde=None, fecha_hasta=None):
    """Genera y retorna bytes del Excel según el tipo de reporte."""
    if tipo == 'ventas':
        datos = reporte_ventas(fecha_desde, fecha_hasta)
        col = ['Fecha', 'Cantidad Pedidos', 'Total Vendido']
        filas = [[d['fecha'], d['cantidad_pedidos'], float(d['total_vendido'])] for d in datos]
        titulo = 'Ventas'
    elif tipo == 'inventario':
        datos = reporte_inventario()
        col = ['Producto', 'Categoría', 'Stock', 'Stock Mín', 'Precio', 'Lote', 'Vencimiento', 'Cant. Lote']
        filas = [[d['pro_nombre'], d['pro_categoria'], d['pro_cantidad_disponible'],
                  d['pro_stock_minimo'], float(d['pro_precio']), d['lot_numero'] or '-',
                  d['lot_fecha_vencimiento'] or '-', d['lot_cantidad_actual'] or 0] for d in datos]
        titulo = 'Inventario'
    elif tipo == 'mas_vendidos':
        datos = reporte_mas_vendidos(fecha_desde, fecha_hasta)
        col = ['Producto', 'Unidades Vendidas', 'Total Vendido']
        filas = [[d['nombre'], d['total_unidades'], float(d['total_vendido'])] for d in datos]
        titulo = 'Mas Vendidos'
    elif tipo == 'por_vencer':
        datos = reporte_por_vencer()
        col = ['Producto', 'Categoría', 'Lote', 'Vencimiento', 'Stock Lote', 'Días Rest.']
        filas = [[d['pro_nombre'], d['pro_categoria'], d['lot_numero'] or d['lot_id'],
                  d['lot_fecha_vencimiento'], d['lot_cantidad_actual'], d['dias_restantes']] for d in datos]
        titulo = 'Por Vencer'
    else:
        return None

    return _generar_excel(titulo, col, filas)
